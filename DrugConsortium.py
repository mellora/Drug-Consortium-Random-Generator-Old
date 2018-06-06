#!/usr/bin/env python3
"""
	Program:		Drug Consortium Random Generator
	License:		MIT License
	Author:			Alexander Mellor
	Description:
		Built with Python 3
		Looks at excel file for company employees
		Looks at config file or user input for number of company randoms
		Pulls a number of employees as set by config file/user input
		Pulls extra employees as alternatives as set by config file/user input
		Outputs employees to both window and text file on desktop
"""
"""
	Imports
"""
# Base with Python
import datetime  # Used to find current date
import os  # Used to find directory paths
import random  # Used seed and randint to calculate the pulls
import tkinter as tk  # Used to create the GUI
from tkinter import Menu
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

# Install with pip
import openpyxl  # Used to parse the excel file

"""
	Variables
"""
mainWindow = tk.Tk()  # Create GUI Instance
random.seed(None)  # Sets random seed with system time
random_pulls = tk.IntVar()  # Initializes variable with int value for number of randoms to pull
random_alternate = tk.IntVar()  # Initializes variable with int value for number of alternates to pull
population_company = []  # Creates master list that will hold company population
population_random = []  # Creates list to hold pulled names for randoms
population_alternate = []  # Creates list to hold pulled names for alternates
population_raw = []  # Creates list to hold the raw company population data
population_pulls = {}  # Creates a dictionary for file save backup on pull
config_settings = []  # Creates list to hold the saved settings that the user saved
radVar = tk.IntVar()  # Variable for Radio buttons
radVar.set(0)  # Sets the index value for radVar
userChoices = ["Input", "Config File"]  # Options for Radio buttons text
path_to_config_file = tk.StringVar()  # Creates a string var for path to company config file
path_to_company_population = tk.StringVar()  # Creates a string var for path to company population file
program_directory = os.path.dirname(
	os.path.realpath(__file__))  # Pulls the directory that this program runs from and puts it into a variable
# Variables used for output save file
today = datetime.datetime.now()
year = today.year
month = today.month
day = today.day
date = str(year) + "-" + str(month) + "-" + str(day)
company_name = ""
pulled_save_file_name = (company_name + "-" + date)
wb = openpyxl.Workbook()

"""
	Constants
"""
SPIN_MAX = 20  # Sets the maximum value for the input
SPIN_MIN = 0  # Sets the minimum value for the input


# ======================
# Functions
# ======================
def license_exists():
	# Checks if the license.txt file exists and creates it if it does not.
	license_text = [
		'MIT License',
		'',
		'Copyright (c) 2018 Alexander Sterling Mellor',
		'',
		'Permission is hereby granted, free of charge, to any person obtaining a copy',
		'of this software and associated documentation files (the "Software"), to deal',
		'in the Software without restriction, including without limitation the rights',
		'to use, copy, modify, merge, publish, distribute, sublicense, and/or sell',
		'copies of the Software, and to permit persons to whom the Software is',
		'furnished to do so, subject to the following conditions:',
		'',
		'The above copyright notice and this permission notice shall be included in all',
		'copies or substantial portions of the Software.',
		'',
		'THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR',
		'IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,',
		'FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE',
		'AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER',
		'LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,',
		'OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE',
		'SOFTWARE.'
	]
	working_directory = os.path.dirname(os.path.realpath(__file__))
	license_path = os.path.join(working_directory, "license.txt")
	if not os.path.isfile(license_path):
		license_file = open(license_path, "w")
		for item in license_text:
			license_file.write("%s\n" % item)
		license_file.close()


def readme_exists():
	# Checks if the readme.txt file exists and creates it if it does not.
	readme_text = [
		'#########################',
		'#      Description      #',
		'#########################',
		'This projects purpose is to randomly pull names from a list that is pulled from',
		'a excel document and return two lists. One list is the list of random names',
		'pulled and the other is of the alternate names pulled.  These lists will be output',
		'within the program and in a file folder on the desktop.',
		'',
		'#########################',
		'#  Running the Program  #',
		'#########################',
		'1)  Select if you want to input the random and alternate pulls or pull from a config file.',
		'2)  Input the values or browse to the config file.',
		'3)  Browse to the excel file that holds the company employee population.',
		'4)  Press the Pull Random Button to calculate the random and alternate pulls.',
		'5)  Copy of the random and alternate pulls is saved to the desktop.',
		'        (Name of the saved file follows this format: company name - today\'s date)',
		'',
		'#########################',
		'#  Save new config file #',
		'#########################',
		'1)  Make sure that the Input radio button is active.',
		'2)  Input the values you want to save.',
		'3)  Go to file > New Company Config to save the config file.',
		'',
		'#########################',
		'#        License        #',
		'#########################',
		'This product is released under the MIT License that is included with the program.'
	]
	working_directory = os.path.dirname(os.path.realpath(__file__))
	readme_path = os.path.join(working_directory, "readme.txt")
	if not os.path.isfile(readme_path):
		readme_path = open(readme_path, "w")
		for item in readme_text:
			readme_path.write("%s\n" % item)
		readme_path.close()


def check_dir_exist():
	# Function checks and makes correct sub directories if they do not exist on program boot
	config_directory = (os.path.join(program_directory, "ConfigFiles"))
	population_directory = (os.path.join(program_directory, "CompanyPopulation"))
	if not os.path.exists(config_directory):
		# Checks if the file path to the directory for the config files exists, if not create it
		os.makedirs(config_directory)
	if not os.path.exists(population_directory):
		# Checks if the file path to the directory for the population files exists, if not create it
		os.makedirs(population_directory)


def reset_seed_and_output():
	# Resets the random seed and clears the output Labels
	random.seed(None)
	output_alternates_population.configure(text="")
	output_random_population.configure(text="")


def user_selection():
	# Function used by Radiobutton in configTab
	rad_sel = radVar.get()
	if rad_sel == 0:
		# Disable Config File Options
		configFileEntry.configure(state="disable")
		configFilePathButton.configure(state="disable")
		# Enable User Input Options
		randomSpin.configure(state="normal")
		alternateSpin.configure(state="normal")
	elif rad_sel == 1:
		# Disable User Input Options
		randomSpin.configure(state="disable")
		alternateSpin.configure(state="disable")
		# Enable Config File Options
		configFileEntry.configure(state="enable")
		configFilePathButton.configure(state="enable")


def open_file_path_to_config():
	# Function that opens a file dialog that retrieves the path to config file
	global path_to_config_file
	path_to_config_file = filedialog.askopenfilename(initialdir=(os.path.join(program_directory, "ConfigFiles")),
	                                                 title="Select Company Config file",
	                                                 filetypes=(("config files", "*.conf"), ("all files", "*.*")))
	configFileEntry.delete(0, tk.END)
	configFileEntry.insert(0, path_to_config_file)


def open_file_path_to_company():
	# Function that opens a file dialog that retrieves the path to company population file
	global path_to_company_population
	global company_name
	path_to_company_population = filedialog.askopenfilename(
		initialdir=(os.path.join(program_directory, "CompanyPopulation")),
		title="Select Company Employee file",
		filetypes=(
			("Company Population Files", "*.xlsx"), ("all files", "*.*")))
	CompanyPopulationEntry.delete(0, tk.END)
	CompanyPopulationEntry.insert(0, path_to_company_population)
	company_name = os.path.splitext(os.path.basename(path_to_company_population))[0]


def clear_lists():
	# Clears the population lists
	global population_company
	global population_random
	global population_alternate
	population_company[:] = []
	population_random[:] = []
	population_alternate[:] = []


def populate_company_population():
	# Populates the population_company list from the excel file
	global population_company
	global wb
	wb = openpyxl.load_workbook(path_to_company_population)
	ws = wb.active
	for row_cells in ws.iter_rows():
		for cell in row_cells:
			population_company.append(cell.value)


def get_random_population():
	# Pulls from the population_company to randomly fill the population_random
	global population_random
	global population_company
	global random_pulls
	for x in range(random_pulls.get()):
		index = random.randint(0, len(population_company) - 1)
		population_random.append(population_company.pop(index))


def get_alternate_population():
	# Pulls from the population_company to randomly fill the population_alternate
	global population_alternate
	global population_company
	global random_alternate
	for x in range(random_alternate.get()):
		index = random.randint(0, len(population_company) - 1)
		population_alternate.append(population_company.pop(index))


def get_config_settings():
	# Opens the config file and pull the config settings
	global config_settings
	global random_pulls
	global random_alternate
	with open(path_to_config_file) as f:
		config_settings = f.read().splitlines()
	random_pulls.set(int(config_settings[0]))
	random_alternate.set(int(config_settings[1]))


def create_save_file():
	# Creates an output file that saves a company's random and alternate pulls in a text document
	global date
	global company_name
	global pulled_save_file_name
	pulled_save_file_name = (company_name + "-" + date + ".txt")
	# Pulls the path to the desktop
	full_path = os.path.join(os.path.expanduser("~"), "Desktop", pulled_save_file_name)
	title_random = "Random Pulls:"
	title_alternate = "Alternate Pulls:"
	test = [title_random]
	for item in population_random:
		test.append(item)
	test.append("")
	test.append(title_alternate)
	for item in population_alternate:
		test.append(item)
	save_file = open(full_path, "w")
	for item in test:
		save_file.write("%s\n" % item)
	save_file.close()


def compute_random_pulls():
	# Uses other functions to calculate the randoms and alternates and outputs the results
	random.seed(None)
	rad_sel = radVar.get()
	clear_lists()
	if rad_sel == 1:
		get_config_settings()
	populate_company_population()
	get_random_population()
	get_alternate_population()
	output_random_population.configure(text=("\n".join(population_random)))
	output_alternates_population.configure(text=("\n".join(population_alternate)))
	tabControl.select(outputTab)
	create_save_file()


def save_config_file():
	# Pulls the values from the spin boxes and creates a config file
	global program_directory
	path = (os.path.join(program_directory, "ConfigFiles"))
	save_config_msg = messagebox.askokcancel("Save Config",
	                                         "Are you sure you want to save these values?\nRandom Pulls: " + str(
		                                         random_pulls.get()) + "\nAlternate Pulls: " + str(
		                                         random_alternate.get()))
	if save_config_msg:
		file_name_path = filedialog.asksaveasfilename(initialdir=path, title="Save Config File",
		                                              filetypes=(("config files", "*.conf"), ("all files", "*.*")))
		config_file = open(file_name_path, "w")
		config_item = [str(random_pulls.get()), str(random_alternate.get())]
		for item in config_item:
			config_file.write("%s\n" % item)
		config_file.close()


def help_message():
	# Shows a message box that sends the user to the documentation for help.
	messagebox.showinfo("Help",
	                    "For in-depth help or instructions for running this program, refer to the readme.txt or user "
	                    "manual.")


def about_message():
	# Shows a message box that tells the user what this program is about
	messagebox.showinfo("About",
	                    "This program pulls both randoms and alternates for random drug testing.  Written in Python "
	                    "3.")


"""
	Code
"""
check_dir_exist()  # Initializes the working directory
license_exists()  # Makes sure license file exists and makes it if it does not.
readme_exists()  # Makes sure the readme file exists and makes it if it does not.

# Add window title
mainWindow.title("Extreme Wellness Random Drug Test Puller")
# Prevent GUI resizing
mainWindow.resizable(False, False)

# ====================
# Creating a Menu Bar
# ====================
menu_bar = Menu(mainWindow)
mainWindow.config(menu=menu_bar)
# Add menu items
# File:
file_menu = Menu(menu_bar, tearoff=0)
file_menu.add_command(label="New Company Config", command=save_config_file)
file_menu.add_separator()
file_menu.add_command(label="Reset", command=reset_seed_and_output)
menu_bar.add_cascade(label="File", menu=file_menu)
# Help:
help_menu = Menu(menu_bar, tearoff=0)
help_menu.add_command(label="Help", command=help_message)
help_menu.add_command(label="About", command=about_message)
menu_bar.add_cascade(label="Help", menu=help_menu)

# ====================
# Create Tabs in mainWindow
# ====================
tabControl = ttk.Notebook(mainWindow)  # Create tabControl
configTab = ttk.Frame(tabControl)  # Create Tab for configuration
tabControl.add(configTab, text="Config")  # Add Tab to tabControl
outputTab = ttk.Frame(tabControl)  # Create Tab for output
tabControl.add(outputTab, text="Output")  # Add tab to tabControl
tabControl.pack(expand=1, fill="both")  # Pack to make visible

# ====================
# Config Tab
# ====================
# Set up master LabelFrame for configTab
configMaster = ttk.LabelFrame(configTab)
configMaster.grid(column=0, row=0, padx=5, pady=5)

# Group items that allow user to select if using user input or config file
configChoose = ttk.LabelFrame(configMaster, text=" Choose User Config or File Config ")
configChoose.grid(column=0, row=0, padx=5, pady=5)
# RadioButtons for selection
for col in range(2):
	chooseRad = tk.Radiobutton(configChoose, text=userChoices[col], variable=radVar, value=col, command=user_selection)
	chooseRad.grid(column=col, row=1, sticky=tk.W)

# Group items dealing with how many random_pulls are used in container
numberOfRandoms = ttk.LabelFrame(configMaster, text=" Number of Randoms Needed ")
numberOfRandoms.grid(column=0, row=1, padx=5, pady=5)
# Create Label For user
randomNumberLabel = ttk.Label(numberOfRandoms, text="Please Input How many Randoms need to be pulled:(Integer)")
randomNumberLabel.grid(column=0, row=0, padx=5, pady=5)
# Create Spinbox for Randoms
randomSpin = tk.Spinbox(numberOfRandoms, from_=SPIN_MIN, to=SPIN_MAX, width=12, textvariable=random_pulls)
randomSpin.grid(column=1, row=0, padx=5, pady=5)

# Group items dealing with how many random_alternates are used in container
numberOfAlternates = ttk.LabelFrame(configMaster, text=" Number of Alternates Needed ")
numberOfAlternates.grid(column=0, row=2, padx=5, pady=5)
# Create Label For user
alternateNumberLabel = ttk.Label(numberOfAlternates,
                                 text="Please Input How many Alternates need to be pulled:(Integer)")
alternateNumberLabel.grid(column=0, row=0, padx=5, pady=5)
# Create Spinbox for Alternates
alternateSpin = tk.Spinbox(numberOfAlternates, from_=SPIN_MIN, to=SPIN_MAX, width=12, textvariable=random_alternate)
alternateSpin.grid(column=1, row=0, padx=5, pady=5)

# Group items that deal with config file path
configFilePathGroup = ttk.LabelFrame(configMaster, text=" Company Config Input ")
configFilePathGroup.grid(column=0, row=3, padx=5, pady=5)
# Create Label for user
configFileLabel = ttk.Label(configFilePathGroup, text="Please input the file path to the config file:")
configFileLabel.grid(column=0, row=0, padx=5, pady=5)
# Create Entry Box for file path to file
configFileEntry = ttk.Entry(configFilePathGroup, width=30, textvariable=path_to_config_file)
configFileEntry.grid(column=0, row=1, padx=5, pady=5)
# Create Button to browse to file
configFilePathButton = ttk.Button(configFilePathGroup, text="Browse", command=open_file_path_to_config)
configFilePathButton.grid(column=1, row=1, padx=5, pady=5)
# Default both the entry box and button to Disabled
configFileEntry.configure(state="disable")
configFilePathButton.configure(state="disable")

# Group items that deal with company employee population file path
companyPopulationFilePath = ttk.LabelFrame(configMaster, text=" Company Employee Input ")
companyPopulationFilePath.grid(column=0, row=4, padx=5, pady=5)
# Create Label for user
CompanyPopulationLabel = ttk.Label(companyPopulationFilePath,
                                   text="Please input the file path to the Company Population file:")
CompanyPopulationLabel.grid(column=0, row=0, padx=5, pady=5)
# Create Entry Box for file path to file
CompanyPopulationEntry = ttk.Entry(companyPopulationFilePath, width=30, textvariable=path_to_company_population)
CompanyPopulationEntry.grid(column=0, row=1, padx=5, pady=5)
# Create Button to browse to file
CompanyPopulationButton = ttk.Button(companyPopulationFilePath, text="Browse", command=open_file_path_to_company)
CompanyPopulationButton.grid(column=1, row=1, padx=5, pady=5)

# Create Button for config input
configButton = ttk.Button(configMaster, text="Pull Randoms", command=compute_random_pulls)
# configButton.grid(column=0, row=2, padx=5, pady=5)
configButton.grid(padx=5, pady=5)

# ====================
# Output Tab
# ====================
# Create Master LabelFrame for outputTab
outputMaster = ttk.LabelFrame(outputTab, text=" Pulled ")
outputMaster.grid(column=0, row=0, padx=5, pady=5)

# Create LabelFrame for Pulled Randoms
outputRandoms = ttk.LabelFrame(outputMaster, text=" Randoms ")
outputRandoms.grid(column=0, row=0, padx=5, pady=5)
# Create Label for random_population output
output_random_population = ttk.Label(outputRandoms, text=("\n".join(population_random)))
output_random_population.grid(column=0, row=0, padx=5, pady=5)

# Create LabelFrame for Pulled Alternates
outputAlternates = ttk.LabelFrame(outputMaster, text=" Alternates ")
outputAlternates.grid(column=1, row=0, padx=5, pady=5)
# Create Label for alternates_population output
output_alternates_population = ttk.Label(outputAlternates, text=("\n".join(population_alternate)))
output_alternates_population.grid(column=0, row=0, padx=5, pady=5)

"""
	Start GUI
"""
mainWindow.mainloop()
